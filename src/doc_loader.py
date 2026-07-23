"""Unified document loader for RAG pipelines.

Supports multiple file formats common in enterprise environments:
- PDF (.pdf)       → PyMuPDF (fitz)
- Word (.docx)     → python-docx
- Excel (.xlsx)    → openpyxl
- Markdown (.md)   → plain text passthrough
- Text (.txt)      → plain text passthrough

Each loader returns a list of ``DocPage`` objects — one per logical page or
sheet. This lets the caller preserve page-level metadata (source, page_num,
doc_type) which becomes Qdrant payload and shows up in citations.

Design Notes
------------
- All loaders are synchronous. Call them with ``asyncio.to_thread`` from
  async handlers to avoid blocking the event loop.
- PyMuPDF is imported lazily so the module can be imported even if PyMuPDF
  is not installed (raises ImportError only on actual use).
- Excel: each sheet becomes a DocPage. Cells are serialised as
  "col1: val1 | col2: val2" rows so the LLM can read tabular data naturally.
- Tables in Word are serialised row-by-row so they survive chunking.

Usage
-----
    from src.doc_loader import load_document

    pages = load_document("invoice.pdf")
    for page in pages:
        print(page.page_num, page.content[:100])
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from loguru import logger


# ---------------------------------------------------------------------------
# Public dataclass
# ---------------------------------------------------------------------------


@dataclass
class DocPage:
    """A single logical page / sheet extracted from a document.

    Attributes
    ----------
    content:
        Plain text content of this page. Already stripped of boilerplate.
    page_num:
        1-indexed page number (PDF) or sheet index (Excel).
        0 for single-page formats (TXT, MD, DOCX with no page breaks).
    source:
        File path or identifier the page came from.
    doc_type:
        One of "pdf", "docx", "xlsx", "markdown", "text".
    metadata:
        Extra key-value pairs: title, author, sheet_name, etc.
    """

    content: str
    page_num: int
    source: str
    doc_type: str
    metadata: dict = field(default_factory=dict)

    def to_dict(self) -> dict:
        return {
            "content": self.content,
            "page_num": self.page_num,
            "source": self.source,
            "doc_type": self.doc_type,
            "metadata": self.metadata,
        }


# ---------------------------------------------------------------------------
# PDF loader — PyMuPDF
# ---------------------------------------------------------------------------


def _load_pdf(path: Path) -> list[DocPage]:
    """Extract text from each page of a PDF using PyMuPDF (fitz).

    Why PyMuPDF over pdfplumber or pypdf?
    - 5-10x faster on large PDFs
    - Better table extraction via page.get_text("blocks")
    - Handles rotated pages, multi-column layouts

    Limitation: scanned PDFs (image-only) return empty text.
    Use an OCR step (pytesseract / EasyOCR) for those.
    """
    try:
        import fitz  # PyMuPDF
    except ImportError as e:
        raise ImportError(
            "PyMuPDF not installed. Run: pip install PyMuPDF"
        ) from e

    pages: list[DocPage] = []

    try:
        doc = fitz.open(str(path))
    except Exception as exc:
        logger.error("PDF open failed", path=str(path), error=str(exc))
        raise

    metadata = {
        "title": doc.metadata.get("title", ""),
        "author": doc.metadata.get("author", ""),
        "total_pages": len(doc),
    }

    for i, page in enumerate(doc, start=1):
        # Extract as plain text (respects reading order)
        text = page.get_text("text").strip()

        # Skip empty pages (common in scanned PDFs)
        if not text:
            logger.debug("PDF page empty (possibly scanned)", page=i, path=str(path))
            continue

        pages.append(
            DocPage(
                content=text,
                page_num=i,
                source=str(path),
                doc_type="pdf",
                metadata={**metadata, "page": i},
            )
        )

    doc.close()
    logger.info("PDF loaded", path=str(path), pages=len(pages))
    return pages


# ---------------------------------------------------------------------------
# Word loader — python-docx
# ---------------------------------------------------------------------------


def _table_to_text(table) -> str:
    """Serialise a docx Table to readable text rows.

    Example output:
        Product | Quantity | Price
        Widget A | 100 | $5.00
    """
    rows = []
    for row in table.rows:
        cells = [cell.text.strip() for cell in row.cells]
        rows.append(" | ".join(cells))
    return "\n".join(rows)


def _load_docx(path: Path) -> list[DocPage]:
    """Extract text + tables from a Word document.

    Returns a single DocPage per document because Word files don't have
    a reliable page concept at the python-docx level (page breaks are
    rendering artefacts). The chunker will split the content later.
    """
    try:
        from docx import Document
        from docx.table import Table
        from docx.text.paragraph import Paragraph
    except ImportError as e:
        raise ImportError(
            "python-docx not installed. Run: pip install python-docx"
        ) from e

    try:
        doc = Document(str(path))
    except Exception as exc:
        logger.error("DOCX open failed", path=str(path), error=str(exc))
        raise

    parts: list[str] = []

    # Walk all block-level elements in document order (paragraphs + tables)
    # This preserves the original reading flow.
    body = doc.element.body
    for child in body:
        tag = child.tag.split("}")[-1] if "}" in child.tag else child.tag
        if tag == "p":
            para = Paragraph(child, doc)
            text = para.text.strip()
            if text:
                style_name = para.style.name if para.style else ""
                if "Heading" in style_name:
                    parts.append(f"\n## {text}\n")
                else:
                    parts.append(text)
        elif tag == "tbl":
            tbl = Table(child, doc)
            table_text = _table_to_text(tbl)
            if table_text.strip():
                parts.append(f"\n[TABLE]\n{table_text}\n[/TABLE]\n")

    content = "\n\n".join(p for p in parts if p.strip())

    core = doc.core_properties
    metadata = {
        "title": core.title or path.stem,
        "author": core.author or "",
        "total_pages": 1,
    }

    logger.info("DOCX loaded", path=str(path), chars=len(content))
    return [
        DocPage(
            content=content,
            page_num=1,
            source=str(path),
            doc_type="docx",
            metadata=metadata,
        )
    ]


# ---------------------------------------------------------------------------
# Excel loader — openpyxl
# ---------------------------------------------------------------------------


def _sheet_to_text(sheet, max_rows: int = 500) -> str:
    """Convert an Excel sheet to readable text.

    Strategy:
    1. First row = header (if it looks like one)
    2. Each data row → "col1: val1 | col2: val2"
    3. Empty rows are skipped

    Why not CSV dump?
    - Column names give the LLM crucial context ("Vendor Name: ABC Corp")
    - Much better retrieval vs raw comma-separated values
    """
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return ""

    # Determine header row
    header = [str(cell) if cell is not None else "" for cell in rows[0]]
    has_header = any(h.strip() for h in header)

    lines: list[str] = []

    if has_header:
        lines.append("Columns: " + " | ".join(h for h in header if h))
        data_rows = rows[1:]
    else:
        data_rows = rows

    for row in data_rows[:max_rows]:
        cells = [str(c) if c is not None else "" for c in row]
        if not any(c.strip() for c in cells):
            continue

        if has_header:
            pairs = [
                f"{h}: {v}"
                for h, v in zip(header, cells)
                if h.strip() and v.strip()
            ]
            lines.append(" | ".join(pairs))
        else:
            lines.append(" | ".join(c for c in cells if c.strip()))

    return "\n".join(lines)


def _load_xlsx(path: Path) -> list[DocPage]:
    """Extract data from each sheet of an Excel workbook.

    Returns one DocPage per sheet. Empty sheets are skipped.
    """
    try:
        import openpyxl
    except ImportError as e:
        raise ImportError(
            "openpyxl not installed. Run: pip install openpyxl"
        ) from e

    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    except Exception as exc:
        logger.error("XLSX open failed", path=str(path), error=str(exc))
        raise

    pages: list[DocPage] = []

    for i, sheet_name in enumerate(wb.sheetnames, start=1):
        sheet = wb[sheet_name]
        content = _sheet_to_text(sheet)

        if not content.strip():
            logger.debug("Excel sheet empty, skipping", sheet=sheet_name)
            continue

        pages.append(
            DocPage(
                content=f"Sheet: {sheet_name}\n\n{content}",
                page_num=i,
                source=str(path),
                doc_type="xlsx",
                metadata={
                    "sheet_name": sheet_name,
                    "total_sheets": len(wb.sheetnames),
                    "title": path.stem,
                },
            )
        )

    wb.close()
    logger.info("XLSX loaded", path=str(path), sheets=len(pages))
    return pages


# ---------------------------------------------------------------------------
# Plain text / Markdown loader
# ---------------------------------------------------------------------------


def _load_text(path: Path, doc_type: str = "text") -> list[DocPage]:
    """Load a plain text or Markdown file as a single DocPage."""
    try:
        content = path.read_text(encoding="utf-8", errors="replace").strip()
    except Exception as exc:
        logger.error("Text file read failed", path=str(path), error=str(exc))
        raise

    if not content:
        return []

    logger.info("Text file loaded", path=str(path), chars=len(content))
    return [
        DocPage(
            content=content,
            page_num=1,
            source=str(path),
            doc_type=doc_type,
            metadata={"title": path.stem},
        )
    ]


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

#: Map file extensions to loader functions
_LOADERS: dict = {
    ".pdf": _load_pdf,
    ".docx": _load_docx,
    ".xlsx": _load_xlsx,
    ".xls": _load_xlsx,
    ".md": lambda p: _load_text(p, "markdown"),
    ".markdown": lambda p: _load_text(p, "markdown"),
    ".txt": lambda p: _load_text(p, "text"),
    ".csv": lambda p: _load_text(p, "text"),
}


def load_document(path: str | Path) -> list[DocPage]:
    """Load a document file and return a list of DocPage objects.

    This is the single entry point for all document types.
    The caller does not need to know the format — the loader dispatches
    based on file extension.

    Parameters
    ----------
    path:
        Absolute or relative path to the document file.

    Returns
    -------
    list[DocPage]
        One entry per logical page / sheet. Empty if the file has no content.

    Raises
    ------
    ValueError
        If the file extension is not supported.
    FileNotFoundError
        If the file does not exist.
    ImportError
        If the required library for the format is not installed.

    Example
    -------
    >>> pages = load_document("report.pdf")
    >>> for p in pages:
    ...     print(f"Page {p.page_num}: {len(p.content)} chars")
    """
    path = Path(path)

    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    ext = path.suffix.lower()
    loader = _LOADERS.get(ext)

    if loader is None:
        supported = ", ".join(sorted(_LOADERS.keys()))
        raise ValueError(
            f"Unsupported file type: '{ext}'. Supported: {supported}"
        )

    logger.info("Loading document", path=str(path), format=ext)
    return loader(path)


def load_document_from_bytes(
    data: bytes,
    filename: str,
) -> list[DocPage]:
    """Load a document from raw bytes (e.g. Gradio file upload).

    Writes to a temp file then calls load_document.
    The temp file is deleted after loading.

    Parameters
    ----------
    data:
        Raw file bytes.
    filename:
        Original filename including extension (used for format detection).
    """
    import tempfile

    suffix = Path(filename).suffix.lower()
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        tmp.write(data)
        tmp_path = Path(tmp.name)

    try:
        pages = load_document(tmp_path)
        for page in pages:
            page.source = filename
        return pages
    finally:
        tmp_path.unlink(missing_ok=True)


def supported_extensions() -> list[str]:
    """Return the list of supported file extensions."""
    return list(_LOADERS.keys())
